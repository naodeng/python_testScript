# -*- coding: utf-8 -*-
import locale
from openpyxl import Workbook
from openpyxl import load_workbook


def new_excel():
    # 创建文件对象
    wb = Workbook()

    # 获取第一个sheet
    ws = wb.active
    # 将数据写入到指定的单元格
    # 写入数字
    ws['A1'] = 1
    # 写入中文
    ws['B1'] = "2"
    ws['C1'] = "3"

    # 写入多个单元格
    ws.append([4, 5, 6])
    ws.append([7, 8, 9])

    # 保存为a.xlsx
    wb.save("a.xlsx")


def read_excel_row():
    # 加载一个已经存在的excel文件
    wb = load_workbook('a.xlsx')
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append(row)

    # print("*" * 30)
    # print(rows)  # 所有行
    # print(rows[0])  # 获取第一行
    # print(rows[0][0])  # 获取第一行第一列的单元格对象
    # print(rows[0][0].value)  # 获取第一行第一列的单元格对象的值

    # print(rows[len(rows) - 1])  # 获取最后行
    # print(rows[len(rows) - 1][len(rows[0]) - 1])  # 获取第后一行和最后一列的单元格对象
    # print(rows[len(rows) - 1][len(rows[0]) - 1].value)  # 获取第后一行和最后一列的单元格对象的值
    # print("*" * 30)
    nrows = ws.max_row
    print('excel所有的列数为' + str(nrows))
    for i in range(0, nrows):
        print(rows[i][0].value)


def read_excel_col():
    wb = load_workbook('a.xlsx')
    ws = wb.active
    cols = []
    cols = []
    for col in ws.iter_cols():
        cols.append(col)
    # print("*" * 30)
    # print(cols)  # 所有列
    # print(cols[0])  # 获取第一列
    # print(cols[0][0])  # 获取第一列的第一行的单元格对象
    # print(cols[0][0].value)  # 获取第一列的第一行的值

    # print("*" * 30)
    # print(cols[len(cols) - 1])  # 获取最后一列
    # print(cols[len(cols) - 1][len(cols[0]) - 1])  # 获取最后一列的最后一行的单元格对象
    # print(cols[len(cols) - 1][len(cols[0]) - 1].value)  # 获取最后一列的最后一行的单元格对象的值
    # print("*" * 30)
    columns = ws.max_column
    print('excel所有的行数为' + str(columns))
    for i in range(0, columns):
        print(cols[i][0].value)


if __name__ == '__main__':
    new_excel()
    read_excel_row()
    read_excel_col()
